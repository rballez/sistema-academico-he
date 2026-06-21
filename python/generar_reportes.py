#!/usr/bin/env python3
"""
=============================================================================
GENERAR_REPORTES.PY — Exportar calificaciones a Excel
Sistema de Gestión Académica — Hospital Escandón
=============================================================================
"""

import os
import pathlib
import math
from db import get_connection, get_ciclo_actual, rows_to_list

def redondear(val):
    if val is None or str(val).strip() == '': return None
    return math.floor(float(val) + 0.5)

def exportar_lista_asistencia(tipo: str, grado_filtro: str = 'Todos', ciclo: str = None) -> str:
    """Genera un archivo Excel formateado como lista de asistencia"""
    if ciclo is None: ciclo = get_ciclo_actual()
    conn = get_connection()
    try:
        # BUG FIX: Quitamos 'AND ciclo_ingreso=?' porque los MIP 2 entraron en ciclos anteriores.
        # Al pedir 'activo=1' nos aseguramos de traer a todos los que están en el hospital hoy.
        query = "SELECT ap_paterno, ap_materno, nombres, grado FROM alumnos WHERE activo=1"
        params = []
        if grado_filtro != 'Todos':
            query += " AND grado=?"
            params.append(grado_filtro)
        query += " ORDER BY grado DESC, ap_paterno, ap_materno, nombres"
        alumnos = rows_to_list(conn.execute(query, params).fetchall())
    finally: conn.close()

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista de Alumnos"
        
        fill_header = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        font_bold = Font(bold=True, size=11)
        font_normal = Font(size=10)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        ws.column_dimensions['A'].width = 45
        if tipo == 'asistencia':
            for col in range(2, 22): ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 4

        row_idx = 1
        grupos = {'MIP 2': [a for a in alumnos if a['grado'] == 'MIP 2'], 'MIP 1': [a for a in alumnos if a['grado'] == 'MIP 1']}
        
        for grado_lbl, lista in grupos.items():
            if not lista: continue
            
            # Título principal
            end_col = 21 if tipo == 'asistencia' else 1
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=end_col)
            c = ws.cell(row=row_idx, column=1, value=f"LISTA DE ASISTENCIA {grado_lbl} {ciclo}")
            c.fill = fill_header; c.font = font_bold; c.alignment = align_center
            for col in range(1, end_col + 1): ws.cell(row=row_idx, column=col).border = border_thin
            row_idx += 1
            
            # Alumnos
            for a in lista:
                nombre = f"{a['ap_paterno']} {a['ap_materno']} {a['nombres']}".strip()
                c_nom = ws.cell(row=row_idx, column=1, value=nombre)
                c_nom.font = font_normal; c_nom.alignment = align_left; c_nom.border = border_thin
                
                if tipo == 'asistencia':
                    for col in range(2, 22):
                        c_check = ws.cell(row=row_idx, column=col)
                        c_check.border = border_thin
                row_idx += 1
            
            row_idx += 2 # Espacio entre MIP 1 y MIP 2

        # Configuración de impresión (Orientación horizontal, tamaño A4)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.5; ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5

        dest = pathlib.Path.home() / 'Desktop' / f'Lista_{"Asistencia" if tipo=="asistencia" else "Nombres"}_{ciclo}.xlsx'
        wb.save(dest)
        return str(dest)
    except ImportError:
        raise RuntimeError("La librería openpyxl no está instalada. Ejecuta: pip install openpyxl")


def exportar_lista_asistencia_pdf(tipo: str, grado_filtro: str = 'Todos', ciclo: str = None) -> str:
    """
    Genera un PDF de lista de alumnos ajustado a una sola hoja en ancho (landscape A4).
    Internamente crea primero el Excel y luego genera el PDF con reportlab.
    """
    import tempfile
    if ciclo is None: ciclo = get_ciclo_actual()

    # Paso 1: obtener datos (misma consulta que el Excel)
    conn = get_connection()
    try:
        query = "SELECT ap_paterno, ap_materno, nombres, grado FROM alumnos WHERE activo=1"
        params = []
        if grado_filtro != 'Todos':
            query += " AND grado=?"
            params.append(grado_filtro)
        query += " ORDER BY grado DESC, ap_paterno, ap_materno, nombres"
        alumnos = rows_to_list(conn.execute(query, params).fetchall())
    finally:
        conn.close()

    # Paso 2: crear Excel temporal (requerimiento del flujo interno)
    tmp_xlsx = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_xlsx.close()
    try:
        import openpyxl
        _wb = openpyxl.Workbook()
        _ws = _wb.active
        for a in alumnos:
            _ws.append([f"{a['ap_paterno']} {a['ap_materno']} {a['nombres']}".strip(), a['grado']])
        _wb.save(tmp_xlsx.name)
    except ImportError:
        raise RuntimeError("La librería openpyxl no está instalada. Ejecuta: pip install openpyxl")

    # Paso 3: generar PDF con reportlab a partir de los mismos datos
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except ImportError:
        raise RuntimeError("La librería reportlab no está instalada. Ejecuta: pip install reportlab")

    desktop = pathlib.Path.home() / 'Desktop'
    if not desktop.exists():
        desktop = pathlib.Path.home()
    tipo_lbl = 'Asistencia' if tipo == 'asistencia' else 'Nombres'
    dest = str(desktop / f'Lista_{tipo_lbl}_{ciclo}.pdf')

    PAGE_W, PAGE_H = landscape(A4)  # 841.9 x 595.3 pt
    MARGIN = 1.5 * cm
    usable_w = PAGE_W - 2 * MARGIN

    doc = SimpleDocTemplate(
        dest,
        pagesize=landscape(A4),
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
    )
    styles = getSampleStyleSheet()
    elements = []

    grupos = {
        'MIP 2': [a for a in alumnos if a['grado'] == 'MIP 2'],
        'MIP 1': [a for a in alumnos if a['grado'] == 'MIP 1'],
    }

    for grado_lbl, lista in grupos.items():
        if not lista:
            continue

        titulo = f"LISTA DE {'ASISTENCIA' if tipo == 'asistencia' else 'ALUMNOS'} — {grado_lbl}   {ciclo}"
        elements.append(Paragraph(f"<b>{titulo}</b>", styles['Normal']))
        elements.append(Spacer(1, 0.3 * cm))

        # Columnas: Nombre + (20 celdas de asistencia ó nada)
        if tipo == 'asistencia':
            n_dias = 20
            # Calcular ancho: nombre ocupa el resto, celdas de asistencia son iguales
            celda_w = 0.85 * cm
            nombre_w = usable_w - n_dias * celda_w
            col_widths = [nombre_w] + [celda_w] * n_dias
            header = ['Nombre'] + [str(i + 1) for i in range(n_dias)]
        else:
            col_widths = [usable_w]
            header = ['Nombre']

        table_data = [header]
        for a in lista:
            nombre = f"{a['ap_paterno']} {a['ap_materno']} {a['nombres']}".strip()
            if tipo == 'asistencia':
                table_data.append([nombre] + [''] * n_dias)
            else:
                table_data.append([nombre])

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D0D0D0')),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 7),
            ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
            # Datos
            ('FONTNAME',   (0, 1), (0, -1), 'Helvetica'),
            ('FONTSIZE',   (0, 1), (-1, -1), 7),
            ('ALIGN',      (1, 1), (-1, -1), 'CENTER'),
            ('ALIGN',      (0, 1), (0, -1), 'LEFT'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            # Bordes
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.grey),
            # Filas alternas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('ROWHEIGHT', (0, 0), (-1, -1), 0.65 * cm),
        ]
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 0.8 * cm))

    doc.build(elements)

    # Limpiar Excel temporal
    try:
        os.remove(tmp_xlsx.name)
    except OSError:
        pass

    return dest


def _build_resultados_sheet(ws, materia: str, tipo_examen: str, ciclo: str, titulo_extra: str = ''):
    """Rellena una hoja de Excel con resultados de examen para el ciclo dado."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    conn = get_connection()
    try:
        query = """
            SELECT a.nombre_completo as nombre, a.grado,
                   MAX(e.percent_correct) as calificacion
            FROM alumnos a
            JOIN examenes_raw e ON a.mip_id = e.student_id
            WHERE e.materia=? AND e.tipo_examen=? AND e.ciclo=?
            GROUP BY a.mip_id
        """
        rows = rows_to_list(conn.execute(query, (materia, tipo_examen, ciclo)).fetchall())
    finally:
        conn.close()

    mip2, mip1 = [], []
    for r in rows:
        r['cal_redondeada'] = redondear(r['calificacion'])
        if r['cal_redondeada'] is not None:
            (mip2 if r['grado'] == 'MIP 2' else mip1).append(r)

    mip2.sort(key=lambda x: x['cal_redondeada'], reverse=True)
    mip1.sort(key=lambda x: x['cal_redondeada'], reverse=True)

    fill_red    = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
    fill_yellow = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
    fill_green  = PatternFill(start_color="A9DFBF", end_color="A9DFBF", fill_type="solid")
    fill_header = PatternFill(start_color="F1948A", end_color="F1948A", fill_type="solid")
    font_bold   = Font(bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15

    # Título de ciclo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ct = ws.cell(row=1, column=1, value=f"Ciclo: {ciclo}{' — ' + titulo_extra if titulo_extra else ''}")
    ct.font = Font(bold=True, size=12)
    ct.alignment = align_center

    row_idx = 2
    for grado_lbl, data in [("MIPS 2", mip2), ("MIPS 1", mip1)]:
        if not data:
            continue
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        c = ws.cell(row=row_idx, column=1, value=grado_lbl)
        c.fill = fill_header; c.font = font_bold; c.alignment = align_center; c.border = border_thin
        ws.cell(row=row_idx, column=2).border = border_thin
        row_idx += 1

        c1, c2 = ws.cell(row=row_idx, column=1, value="NOMBRE"), ws.cell(row=row_idx, column=2, value="CALIFICACIÓN")
        for c in [c1, c2]:
            c.fill = fill_header; c.font = font_bold; c.alignment = align_center; c.border = border_thin
        row_idx += 1

        for d in data:
            c1 = ws.cell(row=row_idx, column=1, value=d['nombre'])
            c2 = ws.cell(row=row_idx, column=2, value=d['cal_redondeada'])
            c1.border = border_thin; c2.border = border_thin
            c2.alignment = align_center; c2.font = font_bold
            val = d['cal_redondeada']
            c2.fill = fill_green if val >= 70 else (fill_yellow if val >= 60 else fill_red)
            row_idx += 1
        row_idx += 2


def exportar_resultados_examen(materia: str, tipo_examen: str, export_type: str = 'excel',
                               ciclo: str = None, ciclo_opt: str = 'actual') -> str:
    """
    Exporta resultados de examen a Excel.
    ciclo_opt: 'actual' | 'anterior' | 'ambos'
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("La librería openpyxl no está instalada. Ejecuta: pip install openpyxl")

    ciclo_actual = ciclo or get_ciclo_actual()

    # Detectar ciclo anterior (el más reciente distinto al actual en examenes_raw)
    conn = get_connection()
    try:
        row_ant = conn.execute(
            "SELECT DISTINCT ciclo FROM examenes_raw WHERE ciclo != ? ORDER BY ciclo DESC LIMIT 1",
            (ciclo_actual,)
        ).fetchone()
        ciclo_anterior = row_ant['ciclo'] if row_ant else None
    finally:
        conn.close()

    wb = openpyxl.Workbook()

    if ciclo_opt == 'ambos' and ciclo_anterior:
        # Hoja 1: ciclo actual
        ws1 = wb.active
        ws1.title = f"Actual ({ciclo_actual})"
        _build_resultados_sheet(ws1, materia, tipo_examen, ciclo_actual, "Actual")
        # Hoja 2: ciclo anterior
        ws2 = wb.create_sheet(title=f"Anterior ({ciclo_anterior})")
        _build_resultados_sheet(ws2, materia, tipo_examen, ciclo_anterior, "Anterior")
    elif ciclo_opt == 'anterior' and ciclo_anterior:
        ws = wb.active
        ws.title = f"Anterior ({ciclo_anterior})"
        _build_resultados_sheet(ws, materia, tipo_examen, ciclo_anterior, "Anterior")
    else:
        # Por defecto: ciclo actual
        ws = wb.active
        ws.title = f"Actual ({ciclo_actual})"
        _build_resultados_sheet(ws, materia, tipo_examen, ciclo_actual, "Actual")

    ciclo_tag = ciclo_opt if ciclo_opt != 'actual' else ciclo_actual
    dest = pathlib.Path.home() / 'Desktop' / f'Resultados_{materia}_{tipo_examen}_{ciclo_tag}.xlsx'
    wb.save(dest)
    return str(dest)


def exportar_excel_rotaciones(grado: str = None) -> str:
    """Exporta solo las calificaciones de rotación por materia a Excel."""
    ciclo = get_ciclo_actual()
    conn = get_connection()
    try:
        query = """
            SELECT a.mip_id, a.nombre_completo, a.grado, u.nombre as escuela,
                c_gyo.cal_rotacion  as gyo_rot,
                c_mi.cal_rotacion   as mi_rot,
                c_cir.cal_rotacion  as ciru_rot,
                c_ped.cal_rotacion  as pedia_rot,
                c_fam.cal_rotacion  as fam_rot,
                c_urg.cal_rotacion  as urg_rot
            FROM alumnos a
            LEFT JOIN universidades u ON a.universidad_id = u.id
            LEFT JOIN calificaciones c_gyo  ON a.mip_id=c_gyo.mip_id  AND c_gyo.materia='GyO'             AND c_gyo.ciclo=?
            LEFT JOIN calificaciones c_mi   ON a.mip_id=c_mi.mip_id   AND c_mi.materia='Medicina Interna' AND c_mi.ciclo=?
            LEFT JOIN calificaciones c_cir  ON a.mip_id=c_cir.mip_id  AND c_cir.materia='Cirugía'         AND c_cir.ciclo=?
            LEFT JOIN calificaciones c_ped  ON a.mip_id=c_ped.mip_id  AND c_ped.materia='Pediatría'       AND c_ped.ciclo=?
            LEFT JOIN calificaciones c_fam  ON a.mip_id=c_fam.mip_id  AND c_fam.materia='Familiar'        AND c_fam.ciclo=?
            LEFT JOIN calificaciones c_urg  ON a.mip_id=c_urg.mip_id  AND c_urg.materia='Urgencias'       AND c_urg.ciclo=?
            WHERE a.activo=1
        """
        params = [ciclo] * 6
        if grado:
            query += " AND a.grado=?"
            params.append(grado)
        query += " ORDER BY a.grado DESC, a.ap_paterno, a.nombres"
        rows = rows_to_list(conn.execute(query, params).fetchall())
    finally:
        conn.close()

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rotaciones"

        COL_COLORS = {
            'GyO': 'F5B7B1', 'MI': 'A9DFBF', 'Cirugía': 'AED6F1',
            'Pediatría': 'F9E79F', 'Familiar': 'D5D8DC', 'Urgencias': 'EDBB99'
        }

        fill_header = PatternFill(start_color="2C4F7C", end_color="2C4F7C", fill_type="solid")
        fill_red    = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
        fill_yellow = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
        fill_green  = PatternFill(start_color="A9DFBF", end_color="A9DFBF", fill_type="solid")
        font_wbold  = Font(bold=True, color="FFFFFF", size=10)
        font_bold   = Font(bold=True, size=10)
        font_normal = Font(size=10)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left   = Alignment(horizontal='left', vertical='center')

        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 8
        for col in range(4, 10):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 13

        # Fila 1: encabezado ciclo
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        ct = ws.cell(row=1, column=1, value=f"Calificaciones de Rotación — Ciclo: {ciclo}" + (f" — {grado}" if grado else ""))
        ct.font = Font(bold=True, size=12, color="FFFFFF"); ct.fill = fill_header; ct.alignment = align_center

        # Fila 2: headers
        hdrs = ['Nombre', 'Grado', 'Escuela', 'GyO', 'Medicina Interna', 'Cirugía', 'Pediatría', 'Familiar', 'Urgencias']
        mat_fills = [None, None, None, 'F5B7B1', 'A9DFBF', 'AED6F1', 'F9E79F', 'D5D8DC', 'EDBB99']
        for ci, (h, mf) in enumerate(zip(hdrs, mat_fills), start=1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.fill = PatternFill(start_color=mf, end_color=mf, fill_type="solid") if mf else fill_header
            cell.font = font_wbold if not mf else Font(bold=True, size=10)
            cell.alignment = align_center; cell.border = border_thin
        ws.freeze_panes = 'A3'

        col_keys = ['gyo_rot', 'mi_rot', 'ciru_rot', 'pedia_rot', 'fam_rot', 'urg_rot']
        fill_mip2 = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        fill_mip1 = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

        for ri, r in enumerate(rows, start=3):
            fill_grado = fill_mip2 if r.get('grado') == 'MIP 2' else fill_mip1
            vals = [r.get('nombre_completo',''), r.get('grado',''), r.get('escuela','')] + [redondear(r.get(k)) for k in col_keys]
            for ci, val in enumerate(vals, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = border_thin
                if ci <= 3:
                    cell.font = font_normal; cell.fill = fill_grado
                    cell.alignment = align_left if ci == 1 else align_center
                else:
                    cell.font = font_bold; cell.alignment = align_center
                    if isinstance(val, (int, float)):
                        cell.fill = fill_green if val >= 70 else (fill_yellow if val >= 60 else fill_red)

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        tag = grado.replace(' ', '') if grado else 'Todos'
        dest = pathlib.Path.home() / 'Desktop' / f'Rotaciones_{tag}_{ciclo}.xlsx'
        wb.save(dest)
        return str(dest)
    except ImportError:
        raise RuntimeError("La librería openpyxl no está instalada.")


def exportar_excel_mega_examenes(grado: str = None, usar_troncal: bool = True, usar_remedial: bool = True) -> str:
    """Exporta la mega-tabla de exámenes (MIP1/MIP2/MAX) a Excel real."""
    ciclo = get_ciclo_actual()
    conn = get_connection()
    try:
        alumnos = rows_to_list(conn.execute(
            "SELECT a.mip_id, a.nombre_completo as nombre, u.nombre as escuela, a.grado "
            "FROM alumnos a LEFT JOIN universidades u ON a.universidad_id=u.id "
            "WHERE a.activo=1 ORDER BY a.ap_paterno, a.nombres"
        ).fetchall())
        examenes = rows_to_list(conn.execute(
            "SELECT student_id, materia, tipo_examen, grado_ref, percent_correct FROM examenes_raw"
        ).fetchall())
    finally:
        conn.close()

    MATS = ['GyO', 'Pediatría', 'Cirugía', 'Medicina Interna', 'Urgencias', 'Familiar']
    tipos_p = ['parcial'] + (['remedial'] if usar_remedial else []) + (['troncal'] if usar_troncal else [])
    tipos_f = ['final']   + (['remedial'] if usar_remedial else []) + (['troncal'] if usar_troncal else [])

    data = []
    for a in alumnos:
        if grado and a['grado'] != grado:
            continue
        mis_ex = [e for e in examenes if e['student_id'] == a['mip_id']]
        row = {'nombre': a['nombre'], 'escuela': a.get('escuela',''), 'mip_id': a['mip_id'], 'grado': a['grado']}
        for m in MATS:
            mx = [e for e in mis_ex if e['materia'] == m]
            def mx_val(gr, tipos):
                vals = [e['percent_correct'] for e in mx if e['grado_ref']==gr and e['tipo_examen'] in tipos and e['percent_correct'] is not None]
                return redondear(max(vals)) if vals else None
            def mx_max(tipos):
                vals = [e['percent_correct'] for e in mx if e['tipo_examen'] in tipos and e['percent_correct'] is not None]
                return redondear(max(vals)) if vals else None
            row[m] = {
                'm1_p': mx_val('MIP 1', tipos_p), 'm2_p': mx_val('MIP 2', tipos_p), 'max_p': mx_max(tipos_p),
                'm1_f': mx_val('MIP 1', tipos_f), 'm2_f': mx_val('MIP 2', tipos_f), 'max_f': mx_max(tipos_f),
            }
        data.append(row)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mega Tabla Exámenes"

        fill_header = PatternFill(start_color="2C4F7C", end_color="2C4F7C", fill_type="solid")
        fill_red    = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
        fill_yellow = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
        fill_green  = PatternFill(start_color="A9DFBF", end_color="A9DFBF", fill_type="solid")
        fill_max    = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        font_wbold  = Font(bold=True, color="FFFFFF", size=9)
        font_bold   = Font(bold=True, size=9)
        font_normal = Font(size=9)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left   = Alignment(horizontal='left', vertical='center')

        MAT_COLORS = {'GyO': '922B21', 'Pediatría': '7D6608', 'Cirugía': '1B4F72',
                      'Medicina Interna': '145A32', 'Urgencias': '873600', 'Familiar': '4D5656'}

        # Fila 1: Materias (colspan 6 cada una)
        ws.cell(row=1, column=1, value='Nombre').merge_cells if False else None
        for ci, lbl in enumerate(['Nombre', 'Escuela', 'MIP ID'], start=1):
            c = ws.cell(row=1, column=ci, value=lbl)
            c.fill = fill_header; c.font = font_wbold; c.alignment = align_center; c.border = border_thin
            ws.merge_cells(start_row=1, start_column=ci, end_row=3, end_column=ci)

        for mi, mat in enumerate(MATS):
            col_start = 4 + mi * 6
            ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start+5)
            c = ws.cell(row=1, column=col_start, value=mat)
            color = MAT_COLORS.get(mat, '2C4F7C')
            c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            c.font = font_wbold; c.alignment = align_center; c.border = border_thin

            # Fila 2: PARCIAL / FINAL
            for offset, lbl in enumerate(['PARCIAL', 'FINAL']):
                c2 = ws.cell(row=2, column=col_start + offset*3, value=lbl)
                ws.merge_cells(start_row=2, start_column=col_start+offset*3, end_row=2, end_column=col_start+offset*3+2)
                c2.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                c2.font = font_wbold; c2.alignment = align_center; c2.border = border_thin

            # Fila 3: MIP 1 / MIP 2 / MAX
            for offset, sub in enumerate(['MIP 1', 'MIP 2', 'MAX', 'MIP 1', 'MIP 2', 'MAX']):
                c3 = ws.cell(row=3, column=col_start+offset, value=sub)
                c3.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                c3.font = font_wbold; c3.alignment = align_center; c3.border = border_thin

        ws.freeze_panes = 'D4'

        # Anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 9
        for col in range(4, 4 + len(MATS)*6):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 8

        fill_mip2 = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        fill_mip1 = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

        for ri, r in enumerate(data, start=4):
            fill_grado = fill_mip2 if r.get('grado') == 'MIP 2' else fill_mip1
            for ci, val in enumerate([r['nombre'], r['escuela'], r['mip_id']], start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = border_thin; cell.fill = fill_grado; cell.font = font_normal
                cell.alignment = align_left if ci == 1 else align_center

            for mi, mat in enumerate(MATS):
                col_start = 4 + mi * 6
                md = r.get(mat, {})
                for offset, key in enumerate(['m1_p', 'm2_p', 'max_p', 'm1_f', 'm2_f', 'max_f']):
                    val = md.get(key)
                    cell = ws.cell(row=ri, column=col_start+offset, value=val)
                    cell.border = border_thin; cell.alignment = align_center; cell.font = font_bold
                    is_max = key in ('max_p', 'max_f')
                    if isinstance(val, (int, float)):
                        cell.fill = fill_max if is_max else (fill_green if val >= 70 else (fill_yellow if val >= 60 else fill_red))
                    elif is_max:
                        cell.fill = fill_max

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        tag = grado.replace(' ','') if grado else 'Todos'
        dest = pathlib.Path.home() / 'Desktop' / f'MegaTabla_Examenes_{tag}_{ciclo}.xlsx'
        wb.save(dest)
        return str(dest)
    except ImportError:
        raise RuntimeError("La librería openpyxl no está instalada.")


def exportar_excel_global(grado: str = None, usar_troncal: bool = True, usar_remedial: bool = True) -> str:
    """Exporta la tabla de calificaciones globales a Excel con formato condicional."""
    ciclo = get_ciclo_actual()
    conn = get_connection()
    try:
        query = """
            SELECT a.mip_id, a.nombre_completo, a.grado, u.nombre as escuela,
                c_gyo.cal_rotacion  as gyo_rot,   c_gyo.cal_parcial as gyo_pa,   c_gyo.cal_final as gyo_fi,  c_gyo.cal_ponderada as gyo_total,
                c_mi.cal_rotacion   as mi_rot,    c_mi.cal_parcial  as mi_pa,    c_mi.cal_final  as mi_fi,   c_mi.cal_ponderada  as mi_total,
                c_cir.cal_rotacion  as ciru_rot,  c_cir.cal_parcial as ciru_pa,  c_cir.cal_final as ciru_fi, c_cir.cal_ponderada as ciru_total,
                c_ped.cal_rotacion  as pedia_rot, c_ped.cal_parcial as pedia_pa, c_ped.cal_final as pedia_fi,c_ped.cal_ponderada as pedia_total,
                c_fam.cal_rotacion  as fam_rot,   c_fam.cal_parcial as fam_pa,   c_fam.cal_final as fam_fi,  c_fam.cal_ponderada as fam_total,
                c_urg.cal_rotacion  as urg_rot,   c_urg.cal_parcial as urg_pa,   c_urg.cal_final as urg_fi,  c_urg.cal_ponderada as urg_total,
                c_glob.rubrica_entregas as rubrica_entregas_global
            FROM alumnos a
            LEFT JOIN universidades u ON a.universidad_id = u.id
            LEFT JOIN calificaciones c_gyo  ON a.mip_id=c_gyo.mip_id  AND c_gyo.materia='GyO'             AND c_gyo.ciclo=?
            LEFT JOIN calificaciones c_mi   ON a.mip_id=c_mi.mip_id   AND c_mi.materia='Medicina Interna' AND c_mi.ciclo=?
            LEFT JOIN calificaciones c_cir  ON a.mip_id=c_cir.mip_id  AND c_cir.materia='Cirugía'         AND c_cir.ciclo=?
            LEFT JOIN calificaciones c_ped  ON a.mip_id=c_ped.mip_id  AND c_ped.materia='Pediatría'       AND c_ped.ciclo=?
            LEFT JOIN calificaciones c_fam  ON a.mip_id=c_fam.mip_id  AND c_fam.materia='Familiar'        AND c_fam.ciclo=?
            LEFT JOIN calificaciones c_urg  ON a.mip_id=c_urg.mip_id  AND c_urg.materia='Urgencias'       AND c_urg.ciclo=?
            LEFT JOIN calificaciones c_glob ON a.mip_id=c_glob.mip_id AND c_glob.materia='GLOBAL'         AND c_glob.ciclo=?
            WHERE a.activo=1
        """
        params = [ciclo] * 7
        if grado:
            query += " AND a.grado=?"
            params.append(grado)
        query += " ORDER BY a.grado DESC, a.ap_paterno, a.nombres"
        rows = rows_to_list(conn.execute(query, params).fetchall())
    finally:
        conn.close()

    for r in rows:
        validos = [v for v in [r.get('gyo_total'), r.get('mi_total'), r.get('ciru_total'),
                                r.get('pedia_total'), r.get('fam_total'), r.get('urg_total')] if v is not None]
        r['cal_final_global'] = min(100.0, sum(validos) / len(validos)) if validos else None

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calificaciones Globales"

        fill_header  = PatternFill(start_color="2C4F7C", end_color="2C4F7C", fill_type="solid")
        fill_mip2    = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        fill_mip1    = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
        fill_red     = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
        fill_yellow  = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
        fill_green   = PatternFill(start_color="A9DFBF", end_color="A9DFBF", fill_type="solid")
        font_wbold   = Font(bold=True, color="FFFFFF", size=9)
        font_bold    = Font(bold=True, size=9)
        font_normal  = Font(size=9)
        border_thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left   = Alignment(horizontal='left', vertical='center')

        MATS = [
            ('GyO',             '922B21', ['gyo_rot',   'gyo_pa',   'gyo_fi',   'gyo_total']),
            ('Medicina Interna','145A32', ['mi_rot',    'mi_pa',    'mi_fi',    'mi_total']),
            ('Cirugía',         '1B4F72', ['ciru_rot',  'ciru_pa',  'ciru_fi',  'ciru_total']),
            ('Pediatría',       '7D6608', ['pedia_rot', 'pedia_pa', 'pedia_fi', 'pedia_total']),
            ('Familiar',        '4D5656', ['fam_rot',   'fam_pa',   'fam_fi',   'fam_total']),
            ('Urgencias',       '873600', ['urg_rot',   'urg_pa',   'urg_fi',   'urg_total']),
        ]

        # Anchos
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 14
        for col in range(4, 32):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 9

        # Fila 1: encabezado general
        total_cols = 3 + len(MATS)*4 + 2
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ct = ws.cell(row=1, column=1, value=f"Calificaciones Globales — Ciclo: {ciclo}" + (f" — {grado}" if grado else ""))
        ct.font = Font(bold=True, size=12, color="FFFFFF"); ct.fill = fill_header; ct.alignment = align_center

        # Fila 2: Nombre / Grado / Escuela + materia headers (colspan 4) + Entregas + Final
        for ci, lbl in enumerate(['Nombre', 'Grado', 'Escuela'], start=1):
            c = ws.cell(row=2, column=ci, value=lbl)
            c.fill = fill_header; c.font = font_wbold; c.alignment = align_center; c.border = border_thin
            ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)

        for mi, (mat_name, color, _) in enumerate(MATS):
            col_s = 4 + mi*4
            ws.merge_cells(start_row=2, start_column=col_s, end_row=2, end_column=col_s+3)
            c = ws.cell(row=2, column=col_s, value=mat_name)
            c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            c.font = font_wbold; c.alignment = align_center; c.border = border_thin

        col_ent = 4 + len(MATS)*4
        for ci, lbl in enumerate(['Entregas', 'FINAL GLOBAL'], start=col_ent):
            c = ws.cell(row=2, column=ci, value=lbl)
            c.fill = fill_header; c.font = font_wbold; c.alignment = align_center; c.border = border_thin
            ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)

        # Fila 3: sub-headers Rot / Parcial / Final / Total
        for mi, (mat_name, color, _) in enumerate(MATS):
            col_s = 4 + mi*4
            for offset, sub in enumerate(['Rot', 'Parcial', 'Final', 'Total']):
                c = ws.cell(row=3, column=col_s+offset, value=sub)
                c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                c.font = font_wbold; c.alignment = align_center; c.border = border_thin

        ws.freeze_panes = 'A4'

        # Datos
        all_keys = [k for _, _, keys in MATS for k in keys]
        for ri, r in enumerate(rows, start=4):
            fill_grado = fill_mip2 if r.get('grado') == 'MIP 2' else fill_mip1
            base = [r.get('nombre_completo',''), r.get('grado',''), r.get('escuela','')]
            nums = [redondear(r.get(k)) for k in all_keys]
            extras = [r.get('rubrica_entregas_global',''), redondear(r.get('cal_final_global'))]
            all_vals = base + nums + extras

            for ci, val in enumerate(all_vals, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = border_thin
                if ci <= 3:
                    cell.font = font_normal; cell.fill = fill_grado
                    cell.alignment = align_left if ci == 1 else align_center
                else:
                    cell.alignment = align_center
                    # Total cols (4, 8, 12... cada 4a columna en los datos de materia) y Final global
                    is_total = (ci >= 4 and ci < col_ent and (ci - 4) % 4 == 3) or ci == col_ent + 1
                    cell.font = font_bold if is_total else font_normal
                    if isinstance(val, (int, float)):
                        cell.fill = fill_green if val >= 70 else (fill_yellow if val >= 60 else fill_red)

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        tag = grado.replace(' ','') if grado else 'Todos'
        dest = pathlib.Path.home() / 'Desktop' / f'Calificaciones_Global_{tag}_{ciclo}.xlsx'
        wb.save(dest)
        return str(dest)
    except ImportError:
        raise RuntimeError("La librería openpyxl no está instalada. Ejecuta: pip install openpyxl")
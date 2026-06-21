#!/usr/bin/env python3
"""
=============================================================================
IMPORTAR_FORMATO_ANTIGUO.PY — Importador de archivos Excel históricos
Sistema de Gestión Académica — Hospital Escandón

ROTACIONES.xlsx (simple):
  Fila 1 = encabezados: Nombre | ID | GyO | MI | CIRUGÍA | PEDIATRÍA | FAMILIAR | URGENCIAS
  Filas 2+: datos.  Valor 0/NP = No presentó (se omite).

CALIFICACIONES.xlsx (compleja, 4 filas de header):
  Fila 1: Nombre | Escuela | ID | CALIFICACIONES (fusionado)
  Fila 2: (vacío x3) | GINECOLOGÍA Y OBSTETRICIA | PEDIATRÍA | CIRUGÍA GENERAL |
           MEDICINA INTERNA | MEDICINA FAMILIAR - URGENCIAS | TRONCAL
  Fila 3: (vacío x3) | PARCIAL | FINAL | PARCIAL | FINAL | ... (por materia)
  Fila 4: (vacío x3) | MIP 1 | MIP 2 | REM | MAX | MIP 1 | MIP 2 | REM | MAX | ...
  Fila 5+: datos.

  Por materia: 8 cols = [P-MIP1, P-MIP2, P-REM, P-MAX, F-MIP1, F-MIP2, F-REM, F-MAX]
  TRONCAL: 1 columna al final (remedial del troncal, afecta todas las materias).
  REM es igual para PARCIAL y FINAL en la misma materia → se importa solo una vez
  como tipo_examen='remedial'.
  MAX se ignora (redundante; se recalcula en el sistema).
  "NP" / None / 0 = No presentó.
=============================================================================
"""

import base64
import io
import re
from db import get_connection, rows_to_list

# Mapeo de nombre de materia en Excel → nombre canónico del sistema
MAPA_MATERIAS = {
    'GINECOLOGÍA Y OBSTETRICIA': 'GyO',
    'GINECOLOGIA Y OBSTETRICIA': 'GyO',
    'GYO': 'GyO',
    'G Y O': 'GyO',
    'GINECO': 'GyO',
    'PEDIATRÍA': 'Pediatría',
    'PEDIATRIA': 'Pediatría',
    'CIRUGÍA GENERAL': 'Cirugía',
    'CIRUGIA GENERAL': 'Cirugía',
    'CIRUGÍA': 'Cirugía',
    'CIRUGIA': 'Cirugía',
    'MEDICINA INTERNA': 'Medicina Interna',
    'MED. INTERNA': 'Medicina Interna',
    'MEDICINA FAMILIAR - URGENCIAS': 'Urg-Fam',   # especial: se divide en 2
    'MEDICINA FAMILIAR-URGENCIAS': 'Urg-Fam',
    'MEDICINA FAMILIAR': 'Familiar',
    'URGENCIAS': 'Urgencias',
    'URG-FAM': 'Urg-Fam',
}

MATERIAS_URG_FAM = ['Urgencias', 'Familiar']


def _leer_xlsx(content_b64: str):
    """Carga workbook de openpyxl desde base64."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl no está instalado.")
    raw = base64.b64decode(content_b64)
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)


def _es_np(val) -> bool:
    """Retorna True si el valor significa 'No Presentó'."""
    if val is None:
        return True
    s = str(val).strip().upper()
    return s in ('NP', '-', '', '0', 'N/A', 'NA') or s == '0'


def _safe_float(val) -> float | None:
    """Convierte a float. Retorna None si es NP/vacío/0."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ('NP', '-', '', 'N/A', 'NA'):
        return None
    try:
        f = float(s.replace(',', '.'))
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def _extraer_id(val) -> str:
    """Extrae un ID de 9 dígitos de un valor."""
    m = re.search(r'\b(\d{9})\b', str(val or '').strip())
    return m.group(1) if m else ''


# =============================================================================
# IMPORTAR ROTACIONES (formato simple: Nombre|ID|GyO|MI|CIRUGÍA|PEDIATRÍA|FAMILIAR|URGENCIAS)
# =============================================================================

ROTACIONES_COL_MAP = {
    3: 'GyO',
    4: 'Medicina Interna',
    5: 'Cirugía',
    6: 'Pediatría',
    7: 'Familiar',
    8: 'Urgencias',
}

# Variantes de encabezados (case-insensitive).
# ORDEN IMPORTANTE: claves más largas / específicas PRIMERO para evitar
# falsos positivos ('MI' está contenido en 'FAMILIAR', 'MEDICINA INTERNA', etc.).
ROTACIONES_HEADER_MAP = {
    # Ginecología (debe ir antes que cualquier substring genérico)
    'GINECOLOGÍA Y OBSTETRICIA': 'GyO', 'GINECOLOGIA Y OBSTETRICIA': 'GyO',
    'GINECOLOGÍA': 'GyO', 'GINECOLOGIA': 'GyO', 'GINECO': 'GyO', 'G Y O': 'GyO', 'GYO': 'GyO',
    # Pediatría
    'PEDIATRÍA': 'Pediatría', 'PEDIATRIA': 'Pediatría', 'PEDIA': 'Pediatría',
    # Cirugía
    'CIRUGÍA GENERAL': 'Cirugía', 'CIRUGIA GENERAL': 'Cirugía',
    'CIRUGÍA': 'Cirugía', 'CIRUGIA': 'Cirugía',
    # Familiar ANTES de 'MI' (porque 'MI' ⊂ 'FAMILIAR')
    'MEDICINA FAMILIAR': 'Familiar', 'FAMILIAR': 'Familiar', 'FAM': 'Familiar',
    # Urgencias ANTES de 'MI'
    'URGENCIAS': 'Urgencias', 'URG': 'Urgencias',
    # Medicina Interna — al final para evitar colisión con substrings
    'MEDICINA INTERNA': 'Medicina Interna', 'MED. INTERNA': 'Medicina Interna', 'MI': 'Medicina Interna',
}


def importar_rotaciones_formato_antiguo(content_b64: str, ciclo_destino: str) -> dict:
    """
    Importa ROTACIONES.xlsx — formato simple.
    Fila 1 = encabezados. Columna 1=Nombre, 2=ID, resto=materias.
    Valor 0 o vacío = No presentó (se omite).
    """
    wb = _leer_xlsx(content_b64)
    ws = wb.active

    rows_iter = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows_iter:
        return {'ok': False, 'error': 'Archivo vacío.'}

    # Detectar fila de encabezados y mapeo col_idx → materia
    header_row_idx = 0
    id_col = 1   # 0-indexed, columna B por defecto
    col_to_materia = {}  # 0-indexed col → materia canónica

    for ri, row in enumerate(rows_iter[:5]):
        for ci, val in enumerate(row):
            s = str(val or '').strip().upper()
            # Detectar columna ID
            if re.search(r'\bID\b|\bMIP\b', s):
                id_col = ci
                header_row_idx = ri
            # Detectar columnas de materia
            canon = None
            for key, mat in ROTACIONES_HEADER_MAP.items():
                if key in s:
                    canon = mat
                    break
            if canon:
                col_to_materia[ci] = canon
        if col_to_materia:
            break

    # Fallback: asumir posiciones fijas (col 0=Nombre, 1=ID, 2=GyO, 3=MI, 4=Ciru, 5=Pedia, 6=Fam, 7=Urg)
    if not col_to_materia:
        col_to_materia = {
            2: 'GyO', 3: 'Medicina Interna', 4: 'Cirugía',
            5: 'Pediatría', 6: 'Familiar', 7: 'Urgencias'
        }
        id_col = 1
        header_row_idx = 0

    conn = get_connection()
    try:
        alumnos_db = {str(r['mip_id']): r for r in rows_to_list(
            conn.execute("SELECT mip_id, grado FROM alumnos WHERE activo=1").fetchall())}

        insertados = 0
        omitidos = 0
        no_encontrados = []

        for row in rows_iter[header_row_idx + 1:]:
            if all(v is None for v in row):
                continue

            id_val = _extraer_id(row[id_col] if id_col < len(row) else '')
            if not id_val:
                omitidos += 1
                continue

            alumno = alumnos_db.get(id_val)
            if not alumno:
                no_encontrados.append(id_val)
                omitidos += 1
                continue

            grado_ref = alumno['grado']

            for ci, materia in col_to_materia.items():
                if ci >= len(row):
                    continue
                val = _safe_float(row[ci])
                if val is None:
                    continue

                val = min(100.0, val)
                conn.execute(
                    """INSERT INTO rotaciones_raw
                       (student_id, earned_points, paper_timestamp, key_version,
                        materia, ciclo, estado)
                       VALUES (?, ?, datetime('now'), ?, ?, ?, 'activo')""",
                    (id_val, val, materia[0].upper(), materia, ciclo_destino)
                )
                insertados += 1

        conn.commit()
        return {
            'ok': True,
            'insertados': insertados,
            'omitidos': omitidos,
            'no_encontrados': no_encontrados[:20],
            'ciclo': ciclo_destino,
        }
    except Exception as e:
        conn.rollback()
        raise
    finally:
        conn.close()


# =============================================================================
# IMPORTAR EXÁMENES (formato complejo: mega-tabla con MIP1/MIP2/MAX por materia)
# =============================================================================
# Estructura de columnas (0-indexed):
#   0=Nombre, 1=Escuela, 2=ID
#   Luego grupos de 6 por materia:
#   [P-MIP1, P-MIP2, P-MAX, F-MIP1, F-MIP2, F-MAX]
#   Fila 2 (0-indexed 1): nombre de materia (fusionado cada 6 cols a partir de col 3)
#   Fila 3 (0-indexed 2): PARCIAL / FINAL (fusionado cada 3)
#   Fila 4 (0-indexed 3): MIP 1 / MIP 2 / MAX
#   Fila 5+ (0-indexed 4+): datos

MATERIAS_TRONCAL = ['Cirugía', 'Medicina Interna', 'Pediatría', 'GyO', 'Urgencias', 'Familiar']


def importar_examenes_formato_antiguo(content_b64: str, ciclo_destino: str) -> dict:
    """
    Importa CALIFICACIONES.xlsx — formato con 8 cols por materia:
    [P-MIP1, P-MIP2, P-REM, P-MAX, F-MIP1, F-MIP2, F-REM, F-MAX] + col TRONCAL al final.

    Orden de materias: GyO, Pediatría, Cirugía General, Medicina Interna,
                       Medicina Familiar-Urgencias, Troncal.

    REM es idéntico en PARCIAL y FINAL para la misma materia → solo se importa una vez
    como tipo_examen='remedial'. MAX se ignora (se recalcula en el sistema).
    TRONCAL se inserta como tipo_examen='troncal' para las 6 materias.
    """
    wb = _leer_xlsx(content_b64)
    ws = wb.active

    rows_list = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows_list) < 5:
        return {'ok': False, 'error': 'El archivo no tiene suficientes filas.'}

    # Detectar filas de encabezado
    header_materia_row = 1
    header_tipo_row    = 2
    header_mip_row     = 3
    data_start_row     = 4

    for ri, row in enumerate(rows_list[:8]):
        row_str = ' '.join(str(v or '').upper() for v in row)
        if any(k in row_str for k in ['GINECOL', 'GINECO', 'PEDIATR', 'CIRUG', 'MEDICINA INTERNA']):
            header_materia_row = ri
            header_tipo_row    = ri + 1
            header_mip_row     = ri + 2
            data_start_row     = ri + 3
            break

    row_mat  = rows_list[header_materia_row]
    row_tipo = rows_list[header_tipo_row] if header_tipo_row < len(rows_list) else []
    row_mip  = rows_list[header_mip_row]  if header_mip_row  < len(rows_list) else []

    # Construir col_map: col_idx → {materia, tipo_examen, grado}
    # tipo_examen: 'parcial' | 'final' | 'remedial'
    # grado: 'MIP 1' | 'MIP 2' | None (para remedial)
    col_map     = {}
    troncal_col = None          # columna del TRONCAL
    rem_seen    = set()         # materias cuyo REM ya fue registrado (evitar duplicados)

    current_materia = None
    current_tipo    = None

    for ci in range(len(row_mat)):
        val_mat  = str(row_mat[ci] or '').strip().upper()
        val_tipo = str(row_tipo[ci] if ci < len(row_tipo) else '').strip().upper()
        val_mip  = str(row_mip[ci]  if ci < len(row_mip)  else '').strip().upper()

        # ── Detectar columna TRONCAL ──────────────────────────────────────────
        if 'TRONCAL' in val_mat:
            troncal_col = ci
            continue

        # ── Actualizar materia si la celda tiene contenido ───────────────────
        if val_mat:
            canon = None
            for key, mat in MAPA_MATERIAS.items():
                if key in val_mat:
                    canon = mat
                    break
            if canon:
                current_materia = canon

        # ── Actualizar tipo (PARCIAL / FINAL) ────────────────────────────────
        if val_tipo in ('PARCIAL', 'FINAL'):
            current_tipo = val_tipo.lower()

        if not current_materia or not current_tipo:
            continue

        # ── Registrar columna según el sub-header MIP ────────────────────────
        if val_mip == 'MIP 1':
            col_map[ci] = {'materia': current_materia, 'tipo': current_tipo, 'grado': 'MIP 1'}

        elif val_mip == 'MIP 2':
            col_map[ci] = {'materia': current_materia, 'tipo': current_tipo, 'grado': 'MIP 2'}

        elif val_mip == 'REM':
            # REM igual en PARCIAL y FINAL para la misma materia → solo una vez
            if current_materia not in rem_seen:
                col_map[ci] = {'materia': current_materia, 'tipo': 'remedial', 'grado': None}
                rem_seen.add(current_materia)

        # MAX: ignorar (redundante, se recalcula)

    if not col_map and troncal_col is None:
        return {'ok': False, 'error': '❌ No se detectaron columnas. Verifica que sea CALIFICACIONES.xlsx.'}

    # Columna ID (por defecto col C = índice 2)
    id_col = 2
    for ci, val in enumerate(rows_list[0]):
        s = str(val or '').strip().upper()
        if re.search(r'\bID\b|\bMIP\b|\bC[OÓ]DIGO\b', s):
            id_col = ci
            break

    conn = get_connection()
    try:
        alumnos_db = {str(r['mip_id']): r for r in rows_to_list(
            conn.execute("SELECT mip_id, grado FROM alumnos WHERE activo=1").fetchall())}

        insertados    = 0
        omitidos      = 0
        no_encontrados = []

        for row in rows_list[data_start_row:]:
            if all(v is None for v in row):
                continue

            id_val = _extraer_id(row[id_col] if id_col < len(row) else '')
            if not id_val:
                omitidos += 1
                continue

            alumno = alumnos_db.get(id_val)
            if not alumno:
                no_encontrados.append(id_val)
                omitidos += 1
                continue

            grado_alumno = alumno['grado']  # 'MIP 1' o 'MIP 2'

            # ── Calificaciones por materia / tipo / grado ─────────────────────
            for ci, info in col_map.items():
                if ci >= len(row):
                    continue

                val = _safe_float(row[ci])
                if val is None:
                    continue
                val = min(100.0, val)

                grado_ref = info['grado'] or grado_alumno  # REM usa el grado actual

                # Urg-Fam se expande a Urgencias + Familiar
                materias_destino = MATERIAS_URG_FAM if info['materia'] == 'Urg-Fam' else [info['materia']]
                for mat in materias_destino:
                    conn.execute(
                        """INSERT INTO examenes_raw
                           (student_id, earned_points, percent_correct, grado_ref,
                            materia, tipo_examen, ciclo)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (id_val, val, val, grado_ref, mat, info['tipo'], ciclo_destino)
                    )
                    insertados += 1

            # ── TRONCAL: insertar una vez para cada materia ───────────────────
            if troncal_col is not None and troncal_col < len(row):
                val_t = _safe_float(row[troncal_col])
                if val_t is not None:
                    val_t = min(100.0, val_t)
                    for mat in MATERIAS_TRONCAL:
                        conn.execute(
                            """INSERT INTO examenes_raw
                               (student_id, earned_points, percent_correct, grado_ref,
                                materia, tipo_examen, ciclo)
                               VALUES (?, ?, ?, ?, ?, ?, ?)""",
                            (id_val, val_t, val_t, grado_alumno, mat, 'troncal', ciclo_destino)
                        )
                        insertados += 1

        conn.commit()
        return {
            'ok': True,
            'insertados': insertados,
            'omitidos': omitidos,
            'no_encontrados': no_encontrados[:20],
            'ciclo': ciclo_destino,
        }
    except Exception as e:
        conn.rollback()
        raise
    finally:
        conn.close()
